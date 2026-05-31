#!/usr/bin/env python3
"""DreamFinder Store Data Converter - mattresses pass (Phase 0 S2).

Reads a completed onboarding workbook (.xlsx) and emits the mattress data files
the live app's build path consumes:

    <output-dir>/data/mattresses.csv        (English - the live CSV contract)
    <output-dir>/data/mattresses-es.csv     (Spanish - only when ES content exists)

and, unless skipped, shells out to <output-dir>/build-data.ps1 to regenerate
<output-dir>/data/mattresses.json from those CSVs (the existing, trusted path).

Usage (run from the repo root, or anywhere):

    python tools/convert_store_data.py <workbook.xlsx> [--output-dir DIR]
           [--build-json | --skip-build-json]

Design (Phase 0 S2, docs/phase0-onboarding-pipeline-spec-2026-05-31.md section3/section4):
  * Mattresses-only. store-config.json, accessories.json, manifest.json, image
    normalization, and allowed-hosts.js are LATER phases (S3-S6) and are NOT
    emitted here. The legacy inline-JS / CSS / footer / --output-html / WebP paths
    have been retired - the live app fetches JSON/config, never inline constants.
  * Header-driven and schema-independent: the Mattresses tab's EN headers ARE the
    live mattresses.csv contract, so we read headers from the workbook and write
    them straight to CSV in workbook order. ES columns are identified purely by a
    trailing " (ES)" suffix. (No dependency on the test workbook_schema module;
    the golden-bundle test guards consistency.)
  * build-data.ps1 is run from <output-dir> (its $PSScriptRoot scopes it to the
    output workspace) so it never touches the repo's data/ unless --output-dir is
    the repo itself (the real onboarding case).

Dependencies: stdlib + openpyxl (already a repo dependency).
"""

from __future__ import annotations

import argparse
import csv
import os
import shutil
import subprocess
import sys

import openpyxl

ES_SUFFIX = " (ES)"


def _to_cell(value) -> str:
    """Workbook cell -> CSV string. Blank for empty; never coerce numbers
    (values originate as text and are passed through verbatim)."""
    if value is None:
        return ""
    return value if isinstance(value, str) else str(value)


def read_mattresses_tab(path):
    """Return (headers, rows). headers in workbook order; rows are dicts
    header->string. Trailing rows without an id are skipped."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if "Mattresses" not in wb.sheetnames:
        wb.close()
        raise SystemExit(f"ERROR: workbook {path!r} has no 'Mattresses' tab "
                         f"(found: {', '.join(wb.sheetnames)})")
    ws = wb["Mattresses"]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [h for h in header_row if h is not None]

    rows = []
    for vals in ws.iter_rows(min_row=2, values_only=True):
        d = {h: _to_cell(vals[i] if i < len(vals) else None)
             for i, h in enumerate(headers)}
        if not d.get("id", "").strip():
            continue  # skip blank/trailing rows
        rows.append(d)
    wb.close()
    return headers, rows


def write_csv(path, fieldnames, rows):
    """Write rows (list of dicts) as CSV. utf-8, newline='' for clean quoting."""
    with open(path, "w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({k: row.get(k, "") for k in fieldnames})


def emit_mattress_csvs(headers, rows, data_dir):
    """Write data/mattresses.csv (EN) and, when ES content exists,
    data/mattresses-es.csv. Returns (en_path, es_path_or_None)."""
    en_headers = [h for h in headers if not h.endswith(ES_SUFFIX)]
    es_headers = [h for h in headers if h.endswith(ES_SUFFIX)]

    # English CSV - EN columns in workbook order (== live CSV contract).
    en_path = os.path.join(data_dir, "mattresses.csv")
    write_csv(en_path, en_headers, rows)

    # Spanish CSV - id + ES columns (suffix stripped). Emit one row per mattress
    # only if any ES cell anywhere is populated; otherwise omit the file entirely.
    es_path = None
    if es_headers:
        es_names = [h[:-len(ES_SUFFIX)] for h in es_headers]
        has_es = any(row.get(h, "").strip() for row in rows for h in es_headers)
        if has_es:
            es_fieldnames = ["id"] + es_names
            es_rows = []
            for row in rows:
                er = {"id": row.get("id", "")}
                for h, name in zip(es_headers, es_names):
                    er[name] = row.get(h, "")
                es_rows.append(er)
            es_path = os.path.join(data_dir, "mattresses-es.csv")
            write_csv(es_path, es_fieldnames, es_rows)

    return en_path, es_path


def run_build_data(output_dir):
    """Invoke <output-dir>/build-data.ps1 to regenerate mattresses.json.
    Warn + skip (never fail the CSV emit) if PowerShell or the script is absent."""
    script = os.path.join(output_dir, "build-data.ps1")
    if not os.path.exists(script):
        print(f"[build-json] skipped: {script} not found.")
        return
    ps = shutil.which("pwsh") or shutil.which("powershell")
    if not ps:
        print("[build-json] skipped: no pwsh/powershell on PATH.")
        return
    print(f"[build-json] running {os.path.basename(ps)} -File {script}")
    proc = subprocess.run(
        [ps, "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", script],
        capture_output=True, text=True)
    if proc.stdout.strip():
        print("  " + proc.stdout.strip().replace("\n", "\n  "))
    if proc.returncode != 0:
        print(f"[build-json] WARNING: build-data.ps1 exited {proc.returncode} "
              f"(CSV output is still valid).")
        if proc.stderr.strip():
            print("  " + proc.stderr.strip().replace("\n", "\n  "))


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Convert an onboarding workbook into DreamFinder mattress CSVs.",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("workbook", help="Path to the onboarding .xlsx")
    parser.add_argument("--output-dir", default=".",
                        help="Where to write data/ (default: current directory)")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--build-json", action="store_true",
                       help="Run build-data.ps1 to regenerate mattresses.json (default).")
    group.add_argument("--skip-build-json", action="store_true",
                       help="Do not invoke build-data.ps1.")
    args = parser.parse_args(argv)

    data_dir = os.path.join(args.output_dir, "data")
    os.makedirs(data_dir, exist_ok=True)

    print(f"Reading {args.workbook} (Mattresses tab)...")
    headers, rows = read_mattresses_tab(args.workbook)
    print(f"  {len(rows)} mattress row(s)")

    en_path, es_path = emit_mattress_csvs(headers, rows, data_dir)
    print(f"  wrote {en_path}")
    if es_path:
        print(f"  wrote {es_path}")
    else:
        print("  no Spanish content - data/mattresses-es.csv omitted")

    if not args.skip_build_json:
        run_build_data(args.output_dir)
    else:
        print("[build-json] skipped (--skip-build-json).")

    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
