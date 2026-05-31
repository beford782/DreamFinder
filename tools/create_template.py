#!/usr/bin/env python3
"""Generate the blank DreamFinder retailer onboarding workbook (.xlsx).

S7 of the Phase 0 plan (docs/phase0-onboarding-pipeline-spec-2026-05-31.md).

The 5 data tabs (Store Info, Brands, Mattresses, Accessories, SalesNotes) are
generated from the SAME shared schema the converter and the Bel fixture use
(tools/workbook_schema.py), so the retailer-facing template can never drift from
what tools/convert_store_data.py reads. A `--self-check` reopens the workbook and
asserts each data tab's row-1 headers equal schema.get_column_headers(tab).

This is a BLANK template (no example rows): the worked, fully-filled example is
the Bel fixture produced by tests/fixtures/build_bel_workbook.py. Two static
helper tabs (Instructions, Feature Keywords) document the pipeline and the
feature-keyword vocabulary.

Usage:
    python tools/create_template.py [--output PATH] [--self-check]

Default output: onboarding/DreamFinder_Onboarding_Template.xlsx (a committed,
retailer-facing deliverable). Dependencies: stdlib + openpyxl. ASCII console
output (workbook cell copy may contain Unicode, but console messages do not).
"""

from __future__ import annotations

import argparse
import os
import sys

# Shared schema lives alongside this file in tools/.
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import workbook_schema as schema  # noqa: E402

import openpyxl  # noqa: E402
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_OUTPUT = os.path.join(REPO_ROOT, "onboarding",
                              "DreamFinder_Onboarding_Template.xlsx")

# -- Styles -------------------------------------------------------------------
HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F3A5C", end_color="1F3A5C", fill_type="solid")
SECTION_FONT = Font(name="Calibri", bold=True, size=12, color="1F3A5C")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F3A5C")
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(left=Side(style="thin"), right=Side(style="thin"),
                     top=Side(style="thin"), bottom=Side(style="thin"))
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER


def set_widths(ws, headers, min_w=12, max_w=45):
    for i, h in enumerate(headers, 1):
        width = min(max(len(str(h)) + 2, min_w), max_w)
        ws.column_dimensions[get_column_letter(i)].width = width


# -- Data validations (enum values hardcoded here; schema carries no enums) -----

def _col(headers, name):
    return get_column_letter(headers.index(name) + 1)


def _add_list(ws, headers, name, options, allow_blank=True):
    if name not in headers:
        return
    dv = DataValidation(type="list", formula1='"' + ",".join(options) + '"',
                        allow_blank=allow_blank)
    ws.add_data_validation(dv)
    c = _col(headers, name)
    dv.add(f"{c}2:{c}1000")


def _add_whole(ws, headers, name, lo, hi):
    if name not in headers:
        return
    dv = DataValidation(type="whole", operator="between", formula1=lo, formula2=hi,
                        allow_blank=True)
    ws.add_data_validation(dv)
    c = _col(headers, name)
    dv.add(f"{c}2:{c}1000")


def apply_validations(ws, tab_name, headers):
    if tab_name == "Mattresses":
        _add_list(ws, headers, "tier", ["gold", "silver", "bronze"], allow_blank=False)
        _add_whole(ws, headers, "firmnessScore", 1, 10)
        _add_list(ws, headers, "locally-made", ["yes", "no"])
    elif tab_name == "Accessories":
        _add_list(ws, headers, "Category",
                  ["Foundations & Support", "Pillows", "Protectors"], allow_blank=False)
        _add_list(ws, headers, "Sub-Type",
                  ["adjustable", "foundation", "low_profile", "bunkie"])
        for h in headers:
            if h.startswith("Score:"):
                _add_whole(ws, headers, h, 0, 5)
    elif tab_name == "SalesNotes":
        _add_list(ws, headers, "Type", ["subBrand", "brand"], allow_blank=False)
        _add_list(ws, headers, "Format", ["full", "coaching"])
    elif tab_name == "Store Info":
        _add_list(ws, headers, "Languages", ["en", "en, es"], allow_blank=False)


# -- Data tabs (schema-driven) ------------------------------------------------

def build_data_tab(wb, tab_name):
    ws = wb.create_sheet(tab_name)
    headers = schema.get_column_headers(tab_name)
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"
    set_widths(ws, headers)
    apply_validations(ws, tab_name, headers)
    return ws


# -- Helper tabs (static; refreshed for the workbook -> bundle pipeline) --------

INSTRUCTIONS = [
    ("DreamFinder Store Onboarding - How to Fill Out This Workbook", "title"),
    ("", ""),
    ("This workbook is converted into your store's DreamFinder data bundle by an", ""),
    ("automated pipeline (workbook -> converter -> bundle). Fill in the data tabs;", ""),
    ("the converter produces store-config.json, mattresses CSV/JSON, accessories.json,", ""),
    ("manifest.json, normalized product images, and the domain-lock allow-list.", ""),
    ("", ""),
    ("Tabs in this workbook:", "section"),
    ("  Store Info    - one data row: store identity, colors, copy (text + voice),", ""),
    ("                  manifest.* fields, allowedHosts, discount settings.", ""),
    ("  Brands        - one row per mattress brand (shown in the footer).", ""),
    ("  Mattresses    - one row per mattress; English columns plus optional (ES) columns.", ""),
    ("  Accessories   - one row per accessory; bilingual name/category/description.", ""),
    ("  SalesNotes    - sub-brand / brand sales notes (full or coaching format).", ""),
    ("", ""),
    ("STORE INFO", "section"),
    ("  Fill the single data row (row 2) under every header.", ""),
    ("  Languages: 'en' or 'en, es'.", ""),
    ("  allowedHosts: your GitHub Pages host, e.g. 'acme.github.io'", ""),
    ("    (localhost is always allowed automatically for previews).", ""),
    ("  manifest.* columns drive manifest.json (the installable-app metadata).", ""),
    ("  Public Asset Root must end with a trailing slash.", ""),
    ("", ""),
    ("MATTRESSES", "section"),
    ("  tier: gold / silver / bronze. firmnessScore: 1-10. locally-made: yes / no.", ""),
    ("  features: pipe-delimited feature keywords (see the Feature Keywords tab) -", ""),
    ("    these drive quiz matching. reason_default is required.", ""),
    ("  (ES) columns are optional Spanish translations; leave blank to fall back to English.", ""),
    ("", ""),
    ("ACCESSORIES", "section"),
    ("  Category: Foundations & Support / Pillows / Protectors.", ""),
    ("  Sub-Type (foundations only): adjustable / foundation / low_profile / bunkie.", ""),
    ("  Score columns are 0-5 (blank = not applicable).", ""),
    ("", ""),
    ("SALESNOTES", "section"),
    ("  Type: subBrand or brand. Format (subBrand only): full or coaching.", ""),
    ("  full -> Lead + Demo + Close. coaching -> RSA Note. brand -> Story.", ""),
    ("  (ES) columns mirror the English fields for Spanish.", ""),
    ("", ""),
    ("IMAGES", "section"),
    ("  Submit product images in any common format; the pipeline normalizes them to", ""),
    ("  optimized JPG. Mattress image files are matched by the lowercased mattress", ""),
    ("  name (e.g. 'Athena' -> athena.jpg). Accessory image files are matched by the", ""),
    ("  file name in the accessory's Image File Name column.", ""),
    ("", ""),
    ("NEED A FILLED EXAMPLE?", "section"),
    ("  The Bel Furniture fixture is the worked, fully-filled example. This template", ""),
    ("  is intentionally blank.", ""),
]


def build_instructions_tab(wb):
    ws = wb.create_sheet("Instructions")
    for r, (text, kind) in enumerate(INSTRUCTIONS, 1):
        cell = ws.cell(row=r, column=1, value=text)
        if kind == "title":
            cell.font = TITLE_FONT
        elif kind == "section":
            cell.font = SECTION_FONT
        else:
            cell.font = BODY_FONT
    ws.column_dimensions["A"].width = 100
    return ws


FEATURE_KEYWORDS = [
    ("plush", "Very soft feel (firmness 1-3)", "Side sleepers, lighter body types"),
    ("medium", "Moderate feel (firmness 4-6)", "Combo sleepers, average body types"),
    ("firm", "Hard feel (firmness 7-10)", "Back/stomach sleepers, heavier body types"),
    ("cooling", "Cooling technology (gel, phase-change, copper)", "Hot sleepers, temperature complaints"),
    ("support", "Strong core support", "Back pain, heavier body types"),
    ("hybrid", "Coils + foam construction", "Most versatile type"),
    ("innerspring", "Traditional coil construction", "Budget-friendly, good support"),
    ("latex", "Contains latex foam layers", "Natural feel, responsive, hypoallergenic"),
    ("responsive", "Quick response to movement", "Combo sleepers, people who move a lot"),
    ("pressure-relief", "Relieves pressure points (hips, shoulders)", "Side sleepers, joint pain"),
    ("motion-isolation", "Minimizes partner movement transfer", "Couples, light sleepers"),
    ("hand-tufting", "Hand-tufted construction (premium)", "Premium / durability seekers"),
    ("copper", "Copper-infused materials", "Antimicrobial, cooling, wellness-focused"),
    ("gel-memory-foam", "Gel-infused memory foam", "Cooling + pressure relief"),
    ("foam-encasement", "Foam-encased edge support", "Edge support, full sleep surface"),
    ("quality", "High-quality materials/construction", "Durability seekers"),
    ("comfort", "General comfort emphasis", "Broad appeal"),
    ("lifetime-warranty", "Comes with lifetime warranty", "Durability / value seekers"),
    ("anti-microbial", "Antimicrobial material properties", "Health-conscious sleepers"),
    ("durability", "Built to last / premium construction", "Long-term investment seekers"),
]


def build_feature_keywords_tab(wb):
    ws = wb.create_sheet("Feature Keywords")
    headers = ["Feature Keyword", "What It Means / When to Use", "Quiz Connection"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header_row(ws, len(headers))
    ws.freeze_panes = "A2"
    for r, (kw, desc, quiz) in enumerate(FEATURE_KEYWORDS, 2):
        ws.cell(row=r, column=1, value=kw).font = Font(name="Calibri", bold=True, size=11)
        ws.cell(row=r, column=2, value=desc).font = BODY_FONT
        ws.cell(row=r, column=3, value=quiz).font = BODY_FONT
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 45
    return ws


# -- Assembly + self-check ----------------------------------------------------

def build_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop default sheet; create tabs in schema order
    for tab in schema.get_tab_names():
        build_data_tab(wb, tab)
    build_instructions_tab(wb)
    build_feature_keywords_tab(wb)
    return wb


def self_check(path):
    """Reopen the workbook; assert each data tab exists with schema headers."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ok = True
    for tab in schema.get_tab_names():
        if tab not in wb.sheetnames:
            print(f"  [FAIL] missing tab: {tab}")
            ok = False
            continue
        header = [c.value for c in next(wb[tab].iter_rows(min_row=1, max_row=1))]
        expected = schema.get_column_headers(tab)
        if header == expected:
            print(f"  [ok]   {tab}: {len(header)} headers match schema")
        else:
            print(f"  [FAIL] {tab}: headers differ from schema")
            ok = False
    wb.close()
    return ok


def main(argv=None) -> int:
    parser = argparse.ArgumentParser(
        description="Generate the blank DreamFinder onboarding workbook from the shared schema.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT,
                        help="Path to write the .xlsx (default: the committed onboarding template).")
    parser.add_argument("--self-check", action="store_true",
                        help="Reopen the workbook and assert tabs/headers match the schema.")
    args = parser.parse_args(argv)

    wb = build_workbook()
    os.makedirs(os.path.dirname(os.path.abspath(args.output)), exist_ok=True)
    wb.save(args.output)
    print(f"Template saved to {args.output}")
    print(f"  data tabs: {', '.join(schema.get_tab_names())}")
    print("  helper tabs: Instructions, Feature Keywords")

    if args.self_check:
        print("Self-check:")
        ok = self_check(args.output)
        print("Self-check:", "PASS" if ok else "FAIL")
        return 0 if ok else 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
