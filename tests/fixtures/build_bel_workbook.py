#!/usr/bin/env python3
"""Generate the Bel golden-bundle onboarding workbook from committed Bel data.

S1b of the Phase 0 plan (docs/phase0-onboarding-pipeline-spec-2026-05-31.md §4).

This reads the *committed* Bel bundle —

    data/store-config.json
    data/mattresses.csv  (+ data/mattresses-es.csv)
    data/accessories.json
    manifest.json

— and emits an .xlsx onboarding workbook whose tabs/columns come entirely from
``tests/fixtures/workbook_schema.py`` (the shared source of truth). The eventual
converter rewrite (S2+) will run on this workbook and should reproduce the same
committed bundle; that is the golden-bundle regression test. This generator is
what makes the fixture *script-generated* (can't silently drift) rather than an
opaque checked-in binary.

Usage (run from the repo root):

    python tests/fixtures/build_bel_workbook.py [--output PATH] [--no-verify]

If --output is omitted the workbook is written to a temp path (printed). The
generated .xlsx is a build artifact and is NOT committed. After writing, a
lightweight round-trip self-check confirms the workbook has the expected tabs,
headers, and per-tab row counts (full semantic deep-equal is canonical.py's job,
later). Exit code is non-zero if the self-check fails.

Dependencies: stdlib + openpyxl (already a repo dependency via create_template.py
and convert_store_data.py). Local/offline only.
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import sys
import tempfile
from pathlib import Path
from typing import Any, Dict, List, Optional

# Make `from tests.fixtures import workbook_schema` resolve regardless of cwd by
# putting the repo root (this file's grandparent's parent) on sys.path first.
REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from tests.fixtures import workbook_schema as schema  # noqa: E402

import openpyxl  # noqa: E402


# ── Source loading ───────────────────────────────────────────────────────────

def load_json(rel_path: str) -> Any:
    with open(REPO_ROOT / rel_path, encoding="utf-8") as f:
        return json.load(f)


def load_csv_rows(rel_path: str) -> List[Dict[str, str]]:
    """Read a CSV into an ordered list of dict rows (utf-8-sig strips any BOM)."""
    path = REPO_ROOT / rel_path
    if not path.exists():
        return []
    with open(path, encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def nested_get(obj: Any, dotted: str) -> Any:
    """Walk a dotted path (e.g. 'colors.storePrimary') through nested dicts."""
    cur = obj
    for part in dotted.split("."):
        if isinstance(cur, dict):
            cur = cur.get(part)
        else:
            return None
    return cur


# ── Per-tab value resolvers (schema column key -> source value) ───────────────

# Store Info friendly (non-dotted) keys -> their dotted path inside store-config.
_STORE_INFO_PATHS = {
    "storeName": "storeName",
    "storeKey": "storeKey",
    "logoMain": "logo.main",
    "logoSub": "logo.sub",
    "colorPrimary": "colors.storePrimary",
    "colorPrimaryLight": "colors.storePrimaryLight",
    "colorPrimaryGlow": "colors.storePrimaryGlow",
    "colorAccent": "colors.accent",
    "gasUrl": "gasUrl",
    "publicAssetRoot": "publicAssetRoot",
    "discountCodePrefix": "discount.codePrefix",
    "discountCodeDigits": "discount.codeDigits",
}
# List-valued config fields rendered as a single comma-separated cell.
_STORE_INFO_LISTS = {"languages", "allowedHosts"}


def store_info_value(key: str, config: dict, manifest: dict) -> Any:
    """Resolve one Store Info column. Manifest keys read manifest.json; text.*/
    voice.* (and _es) read store-config via dotted path; friendly keys use the
    map above; list fields are joined."""
    if key.startswith("manifest."):
        return manifest.get(key[len("manifest."):])
    if key in _STORE_INFO_LISTS:
        vals = config.get(key) or []
        return ", ".join(str(v) for v in vals)
    if key in _STORE_INFO_PATHS:
        return nested_get(config, _STORE_INFO_PATHS[key])
    # text.*, text_es.*, voice.*, voice_es.* — dotted into store-config.
    return nested_get(config, key)


def brand_value(key: str, brand: dict) -> Any:
    if key == "name":
        return brand.get("name", "")
    if key == "logoFile":
        # store-config stores a path (images/brands/restonic.jpg); the workbook
        # column is a bare file name. Emit the basename; the converter re-prefixes.
        logo = brand.get("logo", "")
        return os.path.basename(logo) if logo else ""
    return ""


def mattress_value(key: str, lang: str, row: dict, es_row: dict) -> Any:
    """EN columns read the CSV row by column name (schema key == CSV header).
    ES columns ('<field>_es') read the matched mattresses-es.csv row by '<field>'."""
    if lang == "es":
        es_field = key[:-3] if key.endswith("_es") else key
        return (es_row or {}).get(es_field, "")
    return row.get(key, "")


def accessory_value(key: str, acc: dict) -> Any:
    if key == "matchTags":
        return ", ".join(acc.get("matchTags", []) or [])
    if "." in key:  # name.en, category.es, description.en, matchScores.<k>
        return nested_get(acc, key)
    return acc.get(key)  # id, price, subType, image


def sales_notes_rows(config: dict) -> List[Dict[str, Any]]:
    """Flatten salesNotes (+_es) into row dicts keyed by SalesNotes schema keys.
    Order: all subBrands (config order) then all brands (config order)."""
    notes = config.get("salesNotes", {}) or {}
    notes_es = config.get("salesNotes_es", {}) or {}
    sub = notes.get("subBrands", {}) or {}
    sub_es = notes_es.get("subBrands", {}) or {}
    brands = notes.get("brands", {}) or {}
    brands_es = notes_es.get("brands", {}) or {}

    rows: List[Dict[str, Any]] = []
    for name, entry in sub.items():
        es = sub_es.get(name, {}) or {}
        rows.append({
            "type": "subBrand", "key": name,
            "format": entry.get("format", ""),
            "lead": entry.get("lead", ""), "demo": entry.get("demo", ""),
            "close": entry.get("close", ""), "rsaNote": entry.get("rsaNote", ""),
            "story": "",
            "lead_es": es.get("lead", ""), "demo_es": es.get("demo", ""),
            "close_es": es.get("close", ""), "rsaNote_es": es.get("rsaNote", ""),
            "story_es": "",
        })
    for name, entry in brands.items():
        es = brands_es.get(name, {}) or {}
        rows.append({
            "type": "brand", "key": name, "format": "",
            "lead": "", "demo": "", "close": "", "rsaNote": "",
            "story": entry.get("story", ""),
            "lead_es": "", "demo_es": "", "close_es": "", "rsaNote_es": "",
            "story_es": es.get("story", ""),
        })
    return rows


# ── Workbook assembly ────────────────────────────────────────────────────────

def _write_sheet(wb, tab_name: str, data_rows: List[List[Any]]):
    ws = wb.create_sheet(tab_name)
    ws.append(schema.get_column_headers(tab_name))
    ws.freeze_panes = "A2"  # keep the header visible; no other styling (S1b)
    for row in data_rows:
        ws.append(row)
    return ws


def build_workbook(config: dict, manifest: dict,
                   mattresses: List[dict], es_lookup: Dict[str, dict],
                   accessories: List[dict]):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default empty sheet; we create tabs in order

    # Store Info — single data row.
    si_keys = schema.get_column_keys("Store Info")
    _write_sheet(wb, "Store Info",
                 [[store_info_value(k, config, manifest) for k in si_keys]])

    # Brands — one row per brand, config order.
    b_keys = schema.get_column_keys("Brands")
    _write_sheet(wb, "Brands",
                 [[brand_value(k, b) for k in b_keys]
                  for b in (config.get("brands") or [])])

    # Mattresses — EN + side-by-side ES columns, CSV order. lang per column.
    m_cols = schema.get_columns("Mattresses")
    m_data = []
    for row in mattresses:
        es_row = es_lookup.get((row.get("id") or "").strip(), {})
        m_data.append([mattress_value(c.key, c.lang, row, es_row) for c in m_cols])
    _write_sheet(wb, "Mattresses", m_data)

    # Accessories — one row per accessory, JSON order.
    a_keys = schema.get_column_keys("Accessories")
    _write_sheet(wb, "Accessories",
                 [[accessory_value(k, a) for k in a_keys] for a in accessories])

    # SalesNotes — subBrands then brands.
    sn_keys = schema.get_column_keys("SalesNotes")
    sn_rows = sales_notes_rows(config)
    _write_sheet(wb, "SalesNotes",
                 [[r.get(k, "") for k in sn_keys] for r in sn_rows])

    return wb


# ── Round-trip self-check ────────────────────────────────────────────────────

def self_check(path: str, expected_counts: Dict[str, int]) -> bool:
    """Re-open the workbook and confirm tabs, headers, and row counts.
    Lightweight but meaningful; full deep-equal is canonical.py's job (later)."""
    wb = openpyxl.load_workbook(path, read_only=True)
    ok = True

    expected_tabs = schema.get_tab_names()
    if wb.sheetnames != expected_tabs:
        print(f"  [FAIL] tab order: {wb.sheetnames} != {expected_tabs}")
        ok = False

    for tab in expected_tabs:
        if tab not in wb.sheetnames:
            print(f"  [FAIL] missing tab: {tab}")
            ok = False
            continue
        ws = wb[tab]
        header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        expected_header = schema.get_column_headers(tab)
        if header != expected_header:
            print(f"  [FAIL] {tab} headers differ from schema")
            print(f"         got     : {header}")
            print(f"         expected: {expected_header}")
            ok = False
        data_rows = max(ws.max_row - 1, 0)  # minus header
        want = expected_counts[tab]
        status = "ok" if data_rows == want else "FAIL"
        if data_rows != want:
            ok = False
        print(f"  [{status}] {tab}: {data_rows} data row(s) (expected {want})")

    wb.close()
    return ok


# ── CLI ──────────────────────────────────────────────────────────────────────

def main(argv: Optional[List[str]] = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__,
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--output", "-o", default=None,
                        help="Path to write the .xlsx (default: a temp file, printed).")
    parser.add_argument("--no-verify", action="store_true",
                        help="Skip the round-trip self-check.")
    args = parser.parse_args(argv)

    out_path = args.output or os.path.join(tempfile.gettempdir(), "bel_onboarding.xlsx")

    # Load committed sources.
    config = load_json("data/store-config.json")
    manifest = load_json("manifest.json")
    accessories = load_json("data/accessories.json")
    mattresses = load_csv_rows("data/mattresses.csv")
    es_rows = load_csv_rows("data/mattresses-es.csv")
    es_lookup = {(r.get("id") or "").strip(): r for r in es_rows}

    expected_counts = {
        "Store Info": 1,
        "Brands": len(config.get("brands") or []),
        "Mattresses": len(mattresses),
        "Accessories": len(accessories),
        "SalesNotes": len(sales_notes_rows(config)),
    }

    wb = build_workbook(config, manifest, mattresses, es_lookup, accessories)
    os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
    wb.save(out_path)
    print(f"Wrote workbook: {out_path}")
    print("Expected row counts:", expected_counts)

    if args.no_verify:
        print("Self-check skipped (--no-verify).")
        return 0

    print("Round-trip self-check:")
    ok = self_check(out_path, expected_counts)
    print("Self-check:", "PASS" if ok else "FAIL")
    return 0 if ok else 1


if __name__ == "__main__":
    raise SystemExit(main())
