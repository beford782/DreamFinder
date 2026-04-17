#!/usr/bin/env python3
"""
DreamFinder Store Data Converter
=================================
Reads a completed onboarding spreadsheet (.xlsx) and writes the retailer's
deployment into the config-driven file layout the app reads at runtime:

    data/store-config.json     — branding, colors, text + text_es
    data/mattresses.csv        — mattress lineup (build-data.ps1 generates the JSON)
    data/accessories.json      — accessories with bilingual {en, es} shape
    Code.gs                    — GAS script with retailer strings substituted
    manifest.json              — PWA manifest with retailer store name

Optionally (with --source-images) also converts the retailer's raw mattress and
accessory images into optimized WebP under images/.

Intended to run against a DreamFinder-template clone:

    python tools/convert_store_data.py ./incoming/Acme_Store_Data.xlsx \\
        --image-base-url "https://acme.github.io/DreamFinder" \\
        --source-images ./incoming \\
        --output-dir .

After running, manual steps still required:
  • Drop store-logo / brand logos / PWA icons into images/
  • Create a new GAS deployment and paste the web app URL into index.html
  • Add the new Pages domain to the domain lock array in index.html
  • Run .\\build-data.ps1 to regenerate data/mattresses.json
"""

import argparse
import csv
import glob
import json
import os
import re
import sys

import openpyxl


# ── Image conversion ──────────────────────────────────────────────────────

WEBP_LONG_EDGE = 1000
WEBP_QUALITY = 82
CONVERTIBLE_EXTS = ('.png', '.jpg', '.jpeg', '.PNG', '.JPG', '.JPEG', '.webp', '.WEBP')


def convert_images_to_webp(src_dir, dst_dir, label='images'):
    """Convert every image in src_dir into optimized .webp in dst_dir."""
    try:
        from PIL import Image
    except ImportError:
        print("[!] Pillow not installed — skipping image conversion.")
        print("    Install with: pip install Pillow")
        return []

    if not os.path.isdir(src_dir):
        print(f"[!] {label} source folder not found: {src_dir} — skipping.")
        return []

    os.makedirs(dst_dir, exist_ok=True)
    results = []
    for src_path in sorted(glob.glob(os.path.join(src_dir, '*'))):
        if not src_path.endswith(CONVERTIBLE_EXTS):
            continue
        base = os.path.basename(src_path)
        stem, _ = os.path.splitext(base)
        dst_path = os.path.join(dst_dir, stem.lower() + '.webp')

        try:
            img = Image.open(src_path)
            if img.mode in ('RGBA', 'LA'):
                bg = Image.new('RGB', img.size, (255, 255, 255))
                bg.paste(img, mask=img.split()[-1])
                img = bg
            elif img.mode != 'RGB':
                img = img.convert('RGB')

            w, h = img.size
            if max(w, h) > WEBP_LONG_EDGE:
                if w >= h:
                    new_w = WEBP_LONG_EDGE
                    new_h = int(h * WEBP_LONG_EDGE / w)
                else:
                    new_h = WEBP_LONG_EDGE
                    new_w = int(w * WEBP_LONG_EDGE / h)
                img = img.resize((new_w, new_h), Image.LANCZOS)

            img.save(dst_path, 'WEBP', quality=WEBP_QUALITY, method=6)
            results.append((base, os.path.getsize(src_path), os.path.getsize(dst_path)))
        except Exception as e:
            print(f"[!] Failed to convert {base}: {e}")

    if results:
        before = sum(r[1] for r in results)
        after = sum(r[2] for r in results)
        print(f"  {label}: converted {len(results)} images "
              f"({before/1024/1024:.1f} MB → {after/1024/1024:.1f} MB, "
              f"{100*(1-after/before):.1f}% smaller)")
    return results


# ── Spreadsheet readers ──────────────────────────────────────────────────

def _clean_header(h):
    if not h:
        return h
    return h.replace("*", "").strip().split("\n")[0].strip()


def _is_example_row(vals):
    """Heuristic: detect leftover yellow Bel example rows."""
    joined = " ".join(str(v) for v in vals if v is not None).lower()
    return "bel furniture" in joined or "bel-o-pedic" in joined


def _first_nonempty_row(ws, max_row=10):
    """Return the first data row that isn't empty or a Bel example."""
    for row in ws.iter_rows(min_row=2, max_row=max_row, values_only=True):
        if not any(row):
            continue
        if _is_example_row(row):
            continue
        return row
    # Fallback: whatever is on row 2
    return tuple(c.value for c in ws[2])


def read_store_info(ws):
    headers = [_clean_header(c.value) for c in ws[1]]
    data_row = _first_nonempty_row(ws)

    mapping = {
        "Store Name": "store_name",
        "Logo Line 1": "logo_line1",
        "Logo Line 2": "logo_line2",
        "Primary Color (hex)": "primary_color",
        "Primary Color Light (hex)": "primary_color_light",
        "Trust Signal Text": "trust_signal",
        "Badge Text": "badge_text",
        "Default Location": "default_location",
        "Default Discount %": "default_discount",
        "Email Sender Name": "email_sender",
        "Email Subject Line": "email_subject",
        "Contact Email": "contact_email",
        "Store Phone": "store_phone",
        "Store Address": "store_address",
        "Store Hours": "store_hours",
        "Footer Text": "footer",
        "Social Proof Text": "social_proof",
        "Trust Signal (Spanish)": "trust_signal_es",
        "Badge Text (Spanish)": "badge_text_es",
        "Email Sender Name (Spanish)": "email_sender_es",
        "Email Subject Line (Spanish)": "email_subject_es",
        "Footer Text (Spanish)": "footer_es",
        "Social Proof (Spanish)": "social_proof_es",
        "Enable Spanish Toggle (yes/no)": "enable_spanish",
    }

    info = {}
    for i, header in enumerate(headers):
        if header in mapping and i < len(data_row):
            v = data_row[i]
            info[mapping[header]] = "" if v is None else v
    return info


def read_mattresses(ws):
    """Return list of dicts matching the mattresses.csv schema."""
    headers = [_clean_header(c.value) for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        if _is_example_row(row):
            continue
        vals = dict(zip(headers, row))

        # Firmness → label (used if retailer doesn't provide one)
        firm_raw = vals.get("Firmness (1-10)") or vals.get("Firmness") or 5
        try:
            firmness = int(firm_raw)
        except (TypeError, ValueError):
            firmness = 5
        if firmness <= 3:
            firm_label = "Plush"
        elif firmness <= 6:
            firm_label = "Medium"
        else:
            firm_label = "Firm"

        # Normalize comma-separated spreadsheet values to pipe-separated CSV values
        def pipe(s):
            if not s:
                return ""
            return "|".join(x.strip() for x in str(s).split(",") if x.strip())

        # Features — also map kebab-case keywords straight through (build-data.ps1
        # turns them into camelCase for the runtime features array)
        features = pipe(vals.get("Feature Keywords"))

        locally = str(vals.get("Made Locally (yes/no)") or "").strip().lower()
        if locally not in ("yes", "no"):
            locally = "no"

        rows.append({
            "tier": str(vals.get("Tier") or "bronze").strip().lower(),
            "id": str(vals.get("ID") or "").strip(),
            "name": str(vals.get("Name") or "").strip(),
            "brand": str(vals.get("Brand") or "").strip(),
            "subBrand": str(vals.get("Sub-Brand") or "").strip(),
            "firmnessScore": firmness,
            "firmnessLabel": firm_label,
            "price": "",
            "quizTags": pipe(vals.get("Display Tags")),
            "displayBadges": pipe(vals.get("Display Tags")),
            "highlight": str(vals.get("Highlight") or "").strip(),
            "locally-made": locally,
            "features": features,
            "reason_cooling": str(vals.get("Why: Cooling") or "").strip(),
            "reason_pressureRelief": str(vals.get("Why: Pressure Relief") or "").strip(),
            "reason_motionIsolation": str(vals.get("Why: Motion Isolation") or "").strip(),
            "reason_support": str(vals.get("Why: Support") or "").strip(),
            "reason_plush": str(vals.get("Why: Firmness Feel") or "").strip() if firmness <= 3 else "",
            "reason_medium": str(vals.get("Why: Firmness Feel") or "").strip() if 4 <= firmness <= 6 else "",
            "reason_firm": str(vals.get("Why: Firmness Feel") or "").strip() if firmness >= 7 else "",
            "reason_durability": str(vals.get("Why: Durability") or "").strip(),
            "reason_default": str(vals.get("Why: Default") or "").strip(),
        })
    return rows


MATTRESS_CSV_COLUMNS = [
    "tier", "id", "name", "brand", "subBrand", "firmnessScore", "firmnessLabel",
    "price", "quizTags", "displayBadges", "highlight", "locally-made", "features",
    "reason_cooling", "reason_pressureRelief", "reason_motionIsolation",
    "reason_support", "reason_plush", "reason_medium", "reason_firm",
    "reason_durability", "reason_default",
]


def read_accessories(ws):
    """Return list of accessory dicts matching data/accessories.json shape."""
    headers = [_clean_header(c.value) for c in ws[1]]
    accessories = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        if _is_example_row(row):
            continue
        vals = dict(zip(headers, row))

        acc_id = str(vals.get("ID") or "").strip()
        name_en = str(vals.get("Name") or "").strip()
        name_es = str(vals.get("Name (Spanish)") or "").strip() or name_en
        cat_en = str(vals.get("Category") or "").strip()
        cat_es = str(vals.get("Category (Spanish)") or "").strip() or cat_en
        desc_en = str(vals.get("Description") or "").strip()
        desc_es = str(vals.get("Description (Spanish)") or "").strip() or desc_en

        sub_type = str(vals.get("Sub-Type") or "").strip()

        try:
            price = float(vals.get("Price") or 0)
            if price.is_integer():
                price = int(price)
        except (TypeError, ValueError):
            price = 0

        img_file = str(vals.get("Image File Name") or "").strip()
        if img_file:
            stem, _ = os.path.splitext(img_file)
            image = f"images/accessories/{stem.lower()}.webp"
        else:
            image = ""

        match_tags = [t.strip() for t in str(vals.get("Match Tags") or "").split(",") if t.strip()]

        score_map = {
            "Score: Default": "default",
            "Score: Cooling": "cooling",
            "Score: Hot": "hot",
            "Score: Back Pain": "back_pain",
            "Score: Snoring": "snoring",
            "Score: Premium": "premium",
            "Score: Position Side": "position_side",
            "Score: Position Back": "position_back",
            "Score: Position Stomach": "position_stomach",
        }
        match_scores = {}
        for col, key in score_map.items():
            v = vals.get(col)
            if v is None or v == "" or v == 0:
                continue
            try:
                match_scores[key] = int(v)
            except (TypeError, ValueError):
                continue

        entry = {
            "id": acc_id,
            "name": {"en": name_en, "es": name_es},
            "category": {"en": cat_en, "es": cat_es},
            "price": price,
            "image": image,
            "description": {"en": desc_en, "es": desc_es},
            "matchTags": match_tags,
            "matchScores": match_scores,
        }
        if sub_type:
            entry["subType"] = sub_type
        accessories.append(entry)
    return accessories


def read_brands(ws):
    headers = [_clean_header(c.value) for c in ws[1]]
    brands = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        vals = dict(zip(headers, row))
        name = str(vals.get("Brand Name") or "").strip()
        logo = str(vals.get("Logo File Name") or "").strip()
        if name:
            brands.append({"name": name, "logoFile": logo})
    return brands


# ── Output writers ───────────────────────────────────────────────────────

def hex_to_rgba(hex_str, alpha):
    m = re.match(r"^#?([0-9a-fA-F]{6})$", hex_str or "")
    if not m:
        return f"rgba(139, 26, 26, {alpha})"  # fallback
    r = int(m.group(1)[0:2], 16)
    g = int(m.group(1)[2:4], 16)
    b = int(m.group(1)[4:6], 16)
    return f"rgba({r}, {g}, {b}, {alpha})"


def build_store_config(info, brands):
    """Assemble the data/store-config.json structure."""
    store = info.get("store_name", "") or "Your Store"
    primary = info.get("primary_color", "") or "#8B1A1A"
    primary_light = info.get("primary_color_light", "") or "#a52525"
    glow = hex_to_rgba(primary, 0.15)

    enable_es = str(info.get("enable_spanish") or "yes").strip().lower() == "yes"
    languages = ["en", "es"] if enable_es else ["en"]

    discount_pct = info.get("default_discount") or 5

    text = {
        "pageTitle": f"DreamFinder — {store} Sleep Quiz",
        "metaDescription": f"Take the DreamFinder sleep quiz at {store} and get personalized mattress recommendations.",
        "ogTitle": f"DreamFinder — {store}",
        "trustSignal": info.get("trust_signal") or "",
        "madeBadge": info.get("badge_text") or "",
        "socialProof": info.get("social_proof") or f"Recommended by your {store} sleep team",
        "footer": info.get("footer") or f"&copy; {store}. All rights reserved.",
        "emailPrivacy": f"We'll only use your email to send your results from {store}.",
        "privacyPolicy": "Your information is never shared with third parties.",
        "inStockText": f"In Stock at {store}",
        "emailHeader": f"{store} × DreamFinder",
        "emailSubtext": f"Bring this email to your {store} store",
    }

    text_es = {
        "pageTitle": f"DreamFinder — Prueba de Sueño de {store}",
        "metaDescription": f"Toma la prueba de sueño DreamFinder en {store} y recibe recomendaciones personalizadas de colchones.",
        "ogTitle": f"DreamFinder — {store}",
        "trustSignal": info.get("trust_signal_es") or "",
        "madeBadge": info.get("badge_text_es") or info.get("badge_text") or "",
        "socialProof": info.get("social_proof_es") or f"Confiado por clientes de {store}",
        "footer": info.get("footer_es") or info.get("footer") or f"&copy; {store}. Todos los derechos reservados.",
        "emailPrivacy": "Solo usaremos tu correo para enviarte tus resultados.",
        "privacyPolicy": "Tu información nunca se comparte con terceros.",
        "inStockText": f"Disponible en {store}",
        "emailHeader": f"{store} × DreamFinder",
        "emailSubtext": f"Lleva este correo a tu tienda {store}",
    }

    config = {
        "storeName": store,
        "languages": languages,
        "discountPct": discount_pct,
        "colors": {
            "storePrimary": primary,
            "storePrimaryLight": primary_light,
            "storePrimaryGlow": glow,
        },
        "text": text,
        "text_es": text_es,
    }
    if brands:
        config["brands"] = brands
    return config


def write_mattresses_csv(path, mattress_rows):
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=MATTRESS_CSV_COLUMNS)
        w.writeheader()
        for r in mattress_rows:
            w.writerow({c: r.get(c, "") for c in MATTRESS_CSV_COLUMNS})


def patch_code_gs(src, store_name, email_subject, email_sender):
    """String-substitute Bel references in Code.gs."""
    substitutions = [
        ("Bel Furniture", store_name),
        ("Your DreamFinder Results from Bel Furniture", email_subject or f"Your DreamFinder Results from {store_name}"),
        ("Bel Furniture Sleep Team", email_sender or f"{store_name} Sleep Team"),
        ("Show this email at Bel Furniture to redeem.", f"Show this email at {store_name} to redeem."),
        ("Bel Furniture x DreamFinder", f"{store_name} x DreamFinder"),
        ("Bring this email to your Bel Furniture store", f"Bring this email to your {store_name} store"),
    ]
    out = src
    for before, after in substitutions:
        out = out.replace(before, after)
    return out


def patch_manifest(src, store_name):
    try:
        m = json.loads(src)
    except json.JSONDecodeError:
        return src
    m["name"] = f"DreamFinder — {store_name}"
    m["short_name"] = "DreamFinder"
    m["description"] = f"Personalized sleep consultation for {store_name}"
    return json.dumps(m, indent=2, ensure_ascii=False)


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Convert DreamFinder onboarding spreadsheet into deployment config files")
    parser.add_argument("spreadsheet", help="Path to the .xlsx file")
    parser.add_argument("--image-base-url", default="",
                        help="Base URL where images will be hosted (currently informational; "
                             "app uses relative paths and prepends this at send time via PUBLIC_ASSET_ROOT).")
    parser.add_argument("--output-dir", default=".",
                        help="Directory of the DreamFinder template clone (default: current dir)")
    parser.add_argument("--source-images", default=None,
                        help="Folder containing mattresses/ and accessories/ subfolders of raw images. "
                             "If set, images are auto-converted to optimized WebP into <output-dir>/images/.")
    parser.add_argument("--skip-image-conversion", action="store_true")
    args = parser.parse_args()

    out = os.path.abspath(args.output_dir)
    data_dir = os.path.join(out, "data")
    os.makedirs(data_dir, exist_ok=True)

    print(f"Reading {args.spreadsheet}...")
    wb = openpyxl.load_workbook(args.spreadsheet, data_only=True)

    store_info = read_store_info(wb["Store Info"])
    store_name = store_info.get("store_name") or "Your Store"
    print(f"  Store: {store_name}")

    mattress_rows = read_mattresses(wb["Mattresses"])
    tier_counts = {"gold": 0, "silver": 0, "bronze": 0}
    for m in mattress_rows:
        if m["tier"] in tier_counts:
            tier_counts[m["tier"]] += 1
    print(f"  Mattresses: {len(mattress_rows)} (gold={tier_counts['gold']}, "
          f"silver={tier_counts['silver']}, bronze={tier_counts['bronze']})")

    accessories = read_accessories(wb["Accessories"])
    print(f"  Accessories: {len(accessories)}")

    brands = []
    if "Brands" in wb.sheetnames:
        brands = read_brands(wb["Brands"])
        print(f"  Brands: {len(brands)}")

    if args.source_images and not args.skip_image_conversion:
        print(f"\nConverting source images from {args.source_images}...")
        convert_images_to_webp(
            os.path.join(args.source_images, 'mattresses'),
            os.path.join(out, 'images', 'mattresses'),
            label='mattresses'
        )
        convert_images_to_webp(
            os.path.join(args.source_images, 'accessories'),
            os.path.join(out, 'images', 'accessories'),
            label='accessories'
        )

    # Write store-config.json
    config = build_store_config(store_info, brands)
    cfg_path = os.path.join(data_dir, "store-config.json")
    with open(cfg_path, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2, ensure_ascii=False)
    print(f"\n  Wrote {cfg_path}")

    # Write mattresses.csv
    csv_path = os.path.join(data_dir, "mattresses.csv")
    write_mattresses_csv(csv_path, mattress_rows)
    print(f"  Wrote {csv_path}")

    # Write accessories.json
    acc_path = os.path.join(data_dir, "accessories.json")
    with open(acc_path, "w", encoding="utf-8") as f:
        json.dump(accessories, f, indent=2, ensure_ascii=False)
    print(f"  Wrote {acc_path}")

    # Patch Code.gs in place (if present)
    code_gs_path = os.path.join(out, "Code.gs")
    if os.path.exists(code_gs_path):
        with open(code_gs_path, "r", encoding="utf-8") as f:
            src = f.read()
        patched = patch_code_gs(
            src, store_name,
            store_info.get("email_subject", ""),
            store_info.get("email_sender", "")
        )
        with open(code_gs_path, "w", encoding="utf-8") as f:
            f.write(patched)
        print(f"  Patched {code_gs_path}")

    # Patch manifest.json in place (if present)
    manifest_path = os.path.join(out, "manifest.json")
    if os.path.exists(manifest_path):
        with open(manifest_path, "r", encoding="utf-8") as f:
            src = f.read()
        patched = patch_manifest(src, store_name)
        with open(manifest_path, "w", encoding="utf-8") as f:
            f.write(patched + "\n")
        print(f"  Patched {manifest_path}")

    # Post-run checklist
    pages_url = args.image_base_url or "https://<retailer>.github.io/DreamFinder"
    print("\n" + "=" * 64)
    print("Done. Remaining manual steps:")
    print("=" * 64)
    print("  1. Run .\\build-data.ps1 to regenerate data/mattresses.json")
    print("  2. Drop retailer logos into images/logos/ (store, brand, icons)")
    print("  3. Drop PWA icons into repo root: icon-192.png and icon-512.png")
    print("  4. Create a Google Apps Script deployment:")
    print("     - New Google Sheet → Tools → Apps Script")
    print("     - Paste Code.gs (already patched with retailer strings)")
    print("     - Deploy → New deployment → Web app → Anyone")
    print("     - Paste the /exec URL into index.html's GOOGLE_SCRIPT_URL constant")
    print(f"  5. Add '{_host_from_url(pages_url)}' to the domain lock array in index.html")
    print("  6. Enable GitHub Pages on the new repo (main / root)")
    print("  7. Commit and push")


def _host_from_url(url):
    m = re.match(r"^https?://([^/]+)", url)
    return m.group(1) if m else url


if __name__ == "__main__":
    main()
